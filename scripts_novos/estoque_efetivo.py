import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
import util as util
from util import *
from upload import *
from github import Github

# planilha = pd.ExcelFile('Dados panorama economico cni-iel-daniel.xlsx')
# df1 = pd.read_excel(planilha, 'Sheet1')
# df2 = pd.read_excel(planilha, 'Sheet2')
# sheet_to_df_map = pd.read_excel('Dados panorama economico cni-iel-daniel.xlsx', sheet_name=None)
 
wb = load_workbook(filename = '{path}/Dados panorama economico cni-iel-daniel.xlsx')
sheet_name = wb.sheetnames[2]

print('- Planilha acessada')

ws = wb[sheet_name]
df = pd.DataFrame(ws.values)
df.columns = df.iloc[0]
df = df.iloc[1:]
df.columns = ['x', 'indice']
df['x'] =  pd.to_datetime(df['x'],  format='%d%b%Y:%H:%M:%S.%f')
df.sort_values('x', inplace =True) 
df = df.reset_index(drop=True)
df.dropna(inplace =True)

df['x'] = df['x'].astype(str)
df['x'] = df['x'].str[:7]

serie_estoque_efetivo = []
now = datetime.now()
ano_atual = now.year
ano_anterior = ano_atual - 1

print('- Dados extraídos')

for i, row in df.iterrows():
  if row[0][:4] == str(ano_anterior) or row[0][:4] == str(ano_atual):
    mes = row[0][:7].replace('-', '')
    mes = pd.to_datetime(mes, format='%Y%m', errors='coerce')
    mes = mes.strftime("%Y-%m-%dT%H:%M:%SZ")
    serie_estoque_efetivo.append({'x': mes, 'y': row[1]})

print('- Série criada')

meses_ano = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
ano_referencia = ano_referencia = df['x'][-1:].iloc[0][:4]
mes_referencia = int(df['x'][-1:].iloc[0][-2:])
referencia = meses_ano[mes_referencia-1] + '/' + ano_referencia

valor_cartao = serie_estoque_efetivo[-1:][0]['y']

if valor_cartao < 50:
  direcao = 'down'
else:
  direcao = 'up'

print('- Índice do cartão armazenado')

json_estoque_efetivo  = {
    'nome': 'Estoque efetivo em relação ao planejado da Indústria',
    'descricao': 'O índice varia de 0 a 100. Quanto mais próximo de 0, maior o estoques vendido, mais próximo de 100, maior o estoque mantido.',
    'fonte': 'CNI',
    'stats': [
        {
            'titulo': 'Brasil',
            'valor': valor_cartao,
            'direcao': direcao,
            'desc_serie': 'Número índice mês a mês',
            'serie_tipo': 'data',
            'referencia': referencia,
            'y_label': {
                'prefixo_sufixo': 'sufixo',
                'label': ''
            },
            'serie_labels': {
                'x': 'Data',
                'y': 'Índice'
            },
            'serie': serie_estoque_efetivo,
        },
    ]
}

path_save_json = util.config['path_save_json']['path']
name_json = 'estoque_efetivo'

with open(path_save_json + name_json +'.json', 'w', encoding='utf-8') as f:
    json.dump(json_estoque_efetivo, f, ensure_ascii=False, indent=4)

print('- JSON armazenado')

######################################################
#Upload

upload_files_to_github(name_json)
print('- JSON enviado para o GitHub')