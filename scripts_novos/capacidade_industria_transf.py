import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
import util as util
from util import *
from upload import *
from github import Github


path_planilha = 'G:/IEL/OBSERVATORIO/ATUALIZAÇÃO DE DADOS/Dados panorama economico cni-iel-daniel.xlsx'
num_planilha = 0

wb = load_workbook(filename = path_planilha)
print('- Planilha acessada')
sheet_name = wb.sheetnames[num_planilha]

ws = wb[sheet_name]
df = pd.DataFrame(ws.values)
df.columns = df.iloc[0]
df = df.iloc[1:]
df.columns = ['referencia', 'indice']
df['referencia'] =  pd.to_datetime(df['referencia'],  format='%d%b%Y:%H:%M:%S.%f')
df.sort_values('referencia', inplace =True) 
df = df.reset_index(drop=True)
df.dropna(inplace =True)

df['referencia'] = df['referencia'].astype(str)
df['referencia'] = df['referencia'].str[:7]

serie_capacidade_industria = []
now = datetime.now()
ano_atual = now.year
ano_anterior = ano_atual - 1

print('- Dados extraídos')
for i, row in df.iterrows():
  if row[0][:4] == str(ano_anterior) or row[0][:4] == str(ano_atual):
    mes = row[0][:7].replace('-', '')
    mes = pd.to_datetime(mes, format='%Y%m', errors='coerce')
    mes = mes.strftime("%Y-%m-%dT%H:%M:%SZ")
    valor = round(float(row[1]), 1)
    serie_capacidade_industria.append({'x': mes, 'y': valor})
print('- Série criada')
 
meses_ano = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
ano_referencia = df['referencia'][-1:].iloc[0][:4]
mes_referencia = int(df['referencia'][-1:].iloc[0][-2:])
referencia = meses_ano[mes_referencia-1] + '/' + ano_referencia

valor_periodo_anterior = serie_capacidade_industria[-2:][0]['y']
valor_periodo_atual = serie_capacidade_industria[-2:][1]['y']

# valor_cartao = serie_capacidade_industria[-1:][0]['y']
print('- Índice do cartão armazenado')
if valor_periodo_anterior < valor_periodo_atual:
  direcao_seta = 'up'
elif valor_periodo_atual == valor_periodo_anterior:
  direcao_seta = 'rigth'
else:
  direcao_seta = 'down'

if valor_periodo_atual > 50:
  cor_valor = 'green'
else:
  cor_valor = 'red'

capacidade_json = {
    'nome': 'Utilização da Capacidade Instalada da Indústria',
    'descricao': 'O índice varia de 0 a 100. Valores acima de 50 pontos indicam que a capacidade ficou acima do usual para o mês.',
    'fonte': 'CNI',
    'stats': [
        {
            'titulo': 'Brasil',
            'valor': valor_periodo_atual,
            'direcao': direcao_seta,
            'cor_valor': cor_valor,
            'desc_serie': 'Número índice mês a mês',
            'serie_tipo': 'data',
            'referencia': referencia,
            'y_label': {
                'prefixo_sufixo': 'sufixo',
                'label': ''
            },
            'serie_labels': {
                'x': 'Data',
                'y': 'Índice'
            },
            'serie': serie_capacidade_industria,
        },
    ]
}


path_save_json = util.config['path_save_json']['path']
name_json = 'capacidade_industria'

with open(path_save_json + name_json +'.json', 'w', encoding='utf-8') as f:
    json.dump(capacidade_json, f, ensure_ascii=False, indent=4)

print("- JSON armazenado")

######################################################
#Upload

upload_files_to_github(name_json)
print("- JSON enviado para o GitHub")