import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
import util as util
from util import *
from upload import *
from github import Github


path_planilha= 'G:/IEL/ATENDIMENTO AO CLIENTE WEB 2021/00000 PLANEJAMENTO DESENV EMPRESARIAL 2021/0000 PLANEJAMENTO ESTUDOS E PESQUISAS 2021-IEL-SANDRA/OBSERVATÓRIO/ATUALIZAÇÃO DE DADOS/Dados panorama economico cni-iel-daniel.xlsx'
num_planilha = 3

wb = load_workbook(filename = path_planilha)
print('- Planilha acessada')
sheet_name = wb.sheetnames[num_planilha]
 
ws = wb[sheet_name]
df = pd.DataFrame(ws.values)
df.columns = df.iloc[0]
df = df.iloc[1:]
df.columns = ['referencia', 'indice']
df['referencia'] =  pd.to_datetime(df['referencia'],  format='%d%b%Y:%H:%M:%S.%f')
df.sort_values('referencia', inplace =True) 
df = df.reset_index(drop=True)
df.dropna(inplace =True)

df['referencia'] = df['referencia'].astype(str)
df['referencia'] = df['referencia'].str[:7]

serie_confianca_industrial = []
now = datetime.now()
ano_atual = now.year
ano_anterior = ano_atual - 1

print('- Dados extraídos')

for i, row in df.iterrows():
  if row[0][:4] == str(ano_anterior) or row[0][:4] == str(ano_atual):
    mes = row[0][:7].replace('-', '')
    mes = pd.to_datetime(mes, format='%Y%m', errors='coerce')
    mes = mes.strftime("%Y-%m-%dT%H:%M:%SZ")
    serie_confianca_industrial.append({'x': mes, 'y': round(row[1], 3)})
  # print(row[0][:7])

print('- Série criada')

meses_ano = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
ano_referencia = ano_referencia = df['referencia'][-1:].iloc[0][:4]
mes_referencia = int(df['referencia'][-1:].iloc[0][-2:])
referencia = meses_ano[mes_referencia-1] + '/' + ano_referencia

valor_cartao = serie_confianca_industrial[-1:][0]['y']
print('- Índice do cartão armazenado')
if valor_cartao < 50:
  direcao = 'down'
else:
  direcao = 'up'

json_confianca_industrial  = {
    'nome': 'Índice de Confiança Industrial',
    'descricao': 'O índice varia de 0 a 100. Acima de 50, indica confiança.',
    'fonte': 'CNI',
    'stats': [
        {
            'titulo': 'Brasil',
            'valor': valor_cartao,
            'direcao': direcao,
            'desc_serie': 'Número índice mês a mês',
            'serie_tipo': 'data',
            'referencia': referencia,
            'y_label': {
                'prefixo_sufixo': 'sufixo',
                'label': ''
            },
            'serie_labels': {
                'x': 'Data',
                'y': 'Índice'
            },
            'serie': serie_confianca_industrial,
        },
    ]
}



path_save_json = util.config['path_save_json']['path']
name_json = 'confianca_industrial'

with open(path_save_json + name_json +'.json', 'w', encoding='utf-8') as f:
    json.dump(json_confianca_industrial, f, ensure_ascii=False, indent=4)
print('- JSON armazenado')

######################################################
#Upload

upload_files_to_github(name_json)
print('- JSON enviado para o GitHub')