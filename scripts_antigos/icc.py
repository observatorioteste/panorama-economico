#!/usr/bin/env python
# coding: utf-8

#BIBLIOTECAS
from github import Github
import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import json
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from urllib.request import urlopen
import re
from urllib.request import urlopen, urlretrieve, quote
from urllib.parse import urljoin


################################## DOWNLOAD XLSX ##############################

# Remove the trailing / you had, as that gives a 404 page
url = 'https://www.fecomercio.com.br/pesquisas/indice/icc'
u = urlopen(url)
try:
    html = u.read().decode('latin')
finally:
    u.close()

soup = bs(html, "html.parser")

# Select all A elements with href attributes containing URLs starting with http://
for link in soup.select('a[href^="http://"]'):
    href = link.get('href')
    print(link)

    # Make sure it has one of the correct extensions
    if not any(href.endswith(x) for x in ['.csv','.xls','.xlsx']):
        continue

    filename = href.rsplit('/', 1)[-1]
    print("Downloading %s to %s..." % (href, filename) )
    urlretrieve(href, filename)
    print("Done.")


html_page = urlopen("https://www.fecomercio.com.br/pesquisas/indice/icc")
soup = bs(html_page, "html.parser")
for link in soup.findAll('a'):
    href = link.get('href')
    if any(href.endswith(x) for x in ['.csv','.xls','.xlsx']):
        url = link.get('href')

##################################  ##############################
r = requests.get(url, allow_redirects=True)
local_path = 'G:/IEL/ATENDIMENTO AO CLIENTE WEB 2020/00000 PLANEJAMENTO DESENV EMPRESARIAL 2020/00003 PLANEJAMENTO ESTUDOS E PESQUISAS 2020-IEL-SANDRA/OBSERVATÓRIO/OBSERVATÓRIO FIEG HOME ESPACE/arquivos scripts wendel/ICC/icc_link_download_202009.xlsx'
open(local_path, 'wb').write(r.content)
wb = load_workbook(filename = local_path)
sheet_name = wb.sheetnames[1]

ws = wb[sheet_name]
data = pd.DataFrame(ws.values)

# data = data.rename(columns=data.iloc[0]).drop(data.index[0])
data = data[1:]
data.rename(columns=data.iloc[0], inplace=True)
data = data.rename(columns=data.iloc[0]).drop(data.index[0])
data =  data[['Mês', 'ICC ']] 
data = data.loc[data['Mês'] >= datetime(2019, 1, 1, 0, 0)]

################################## INDICE E SERIE ##############################

# #INDICE
data['Ano'] = pd.DatetimeIndex(data['Mês']).year

anos = np.unique(data['Ano'])
mask = (data['Ano'] == anos[len(anos)-1]) 
data_ano_atual = data.loc[mask]


anos = np.unique(data['Ano'])
mask = (data['Ano'] == anos[len(anos)-2]) 
data_ano_anterior = data.loc[mask][:len(data_ano_atual)]

media_ano_anterior = data_ano_anterior['ICC '].mean()
media_ano_atual = data_ano_atual['ICC '].mean()

variacao = (media_ano_atual/media_ano_anterior - 1)*100
valor_cartao = round(variacao, 2)

#SERIE 
data.reset_index(drop=True, inplace=True)
data = data[['Mês', 'ICC ', 'Ano']]
valor = data['ICC '] 
mes = data['Mês'][1].strftime('%Y%m')
mes
serie_icc = []

for row in data.iterrows():
  serie_icc.append({'mes': row[1][0].strftime('%Y%m'), 'valor': row[1][1]})

# # #################################################
# #Referencia
data['Mês_Num'] = pd.DatetimeIndex(data['Mês']).month
mes = list(data['Mês_Num'][-1:])[0]
ano = list(data['Ano'][-1:])[0]
meses_ano = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
referencia_extenso = meses_ano[mes-1] + '/' + str(ano)
referencia = str(mes) + '/' + str(ano)

# #################################################
#Criação do JSON

json_icc = {
    'nome_indice': 'ICC - Índice de Confiança do Consumidor',
    'fonte': 'FecomercioSP - Federação do Comércio de Bens, Serviços e Turismo do Estado de São Paulo', 
    'referencia': referencia, 
    'referencia_extenso': referencia_extenso,
    'valor_cartao': valor_cartao,
    'descricao_cartao': 'Variação em relação ao mesmo período do ano anterior',
    'descricao_serie': 'Variação mensal',
    'serie_icc': serie_icc,
    'periodicidade': 30
}


with open('C:/Users/wendelsouza.iel/Desktop/panorama-economico/icc.json', 'w', encoding='utf-8') as f:
    json.dump(json_icc, f, ensure_ascii=False, indent=4)


######################################################
#Upload
g = Github("observatorioteste", "a6beae056fd437a3ceb2bd23f1c4ce8ceb46fe0d")

repo = g.get_user().get_repo('automatizacao_panorama_economico')

all_files = []
contents = repo.get_contents("")
while contents:
    file_content = contents.pop(0)
    if file_content.type == "dir":
        contents.extend(repo.get_contents(file_content.path))
    else:
        file = file_content
        all_files.append(str(file).replace('ContentFile(path="','').replace('")',''))


with open('C:/Users/wendelsouza.iel/Desktop/panorama-economico/icc.json', 'r', encoding='utf-8') as file:
    content = file.read()
    
# content = content.encode("windows-1252").decode("utf-8")


# Upload to github
git_file = 'panorama-economico/icc.json'
if git_file in all_files:
    contents = repo.get_contents(git_file)
    repo.update_file(contents.path, "committing files", content, contents.sha, branch="observatorioteste-patch-1")
    print(git_file + ' ATUALIZADO!')
else:
    repo.create_file(git_file, "committing files", content, branch="observatorioteste-patch-1")
    print(git_file + ' CRIADO!')